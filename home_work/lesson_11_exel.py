import openpyxl

name_of_fields = ["ім'я", "Клас", "вік"]
fields_01 = ["ЄВГЕН", "2", "10"]
fields_02 = ["Сашко", "5", "12"]
fields_03 = ["Юра", "6", "13"]

wb = openpyxl.Workbook()                  # створюэм об*ект книжка
wb.create_sheet(title='first', index=0)   # створюємо сторінку
print(wb.sheetnames)                      # список створених аркушів
wb.remove(wb['Sheet'])
print(wb.sheetnames)                      # видалення аркуша

sheet = wb['first']

for row_index, row in enumerate((name_of_fields, fields_01, fields_02, fields_03)): # розбираємо рядки
    for call_index, value in enumerate(row):                                        # розбираємо сам рядок
        cell = sheet.cell(row=row_index +1, column=call_index + 1)         # обираэмо ячейку з координатами
        cell.value = value                                                 # записуємо в ячейку значення

wb.save("task_04.xlsx")

#-------------------------------- ВІДКРИТИ ФАЙЛ ЕХЕL--------------------------------------------------------------
wb = openpyxl.load_workbook('task_04.xlsx')                                # загружаємо книгу
sheet = wb.active                                                          #  обираємо активний лист
print(sheet['A1'].value)                                                   # стукаємось до ячейки

rows = sheet.max_row                                                     #  максимально  рядоків
cols = sheet.max_column                                                   #  максимально стовпчиків
print(rows)
print(cols)

name_of_field = []                             #  список назв стопчиків
data = []
for i in range(1,  rows + 1):
    field = []
    for j in range(1, cols + 1):
        cell = sheet.cell(row=i, column=j)
        if i == 1:                             #  якшо і перша строка  то додаємо дані
            name_of_field.append(cell.value)
        else:                                  #  якшо не перша строка
            field.append(cell.value)
    if field:                                  # якщо у полі щось є додоаємо до нашого списку  і він обновляєтсья
        data.append(field)

print(name_of_field)                           #  назви стовпчиків
print(data)                                    #  дані