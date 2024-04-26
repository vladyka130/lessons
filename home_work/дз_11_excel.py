#Прочитати збережений csv-файл із попереднього завдання та зберегти дані в excel-файл.
#Крім віку - стовпець з цими даними не потрібний.

#*Додаткове завдання не обов'язкове до виконання:
#розгорнути таблицю на дев'яносто градусів (стовпці стають рядками, а рядки стовпцями).
#До завдання прикріплений приклад як має виглядати змісту підсумкового файлу.

import openpyxl

with open('my_first.csv') as file:
    reader_csv = [[el.strip()] for el in file]

wb = openpyxl.Workbook()
wb.create_sheet('first_list', index=0)
wb.remove(wb['Sheet'])
print(wb.sheetnames)
my_sheet = wb['first_list']

index, ind, col = 0, 1, 1

for el in reader_csv:
    i = [k.split(',') for k in el]
    for k in i:
        for i in k:
            cell = my_sheet.cell(row=ind, column=col)
            cell.value = i
            col +=1
        col = 1
        ind +=1

my_sheet.delete_cols(3)
wb.save('home_work_11_3.xlsx')

# у додатковому завданні у рядку 30 міняємо місцями  'ind` та `col` для аргументів  row та column





